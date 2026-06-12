from flask import Flask, request, jsonify
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from io import BytesIO
import base64
import math

app = Flask(__name__)

def safe_num(value, default=0):
    if value is None:
        return default
    if isinstance(value, dict):
        return default
    try:
        v = float(value)
        return default if math.isnan(v) or math.isinf(v) else v
    except (TypeError, ValueError):
        return default

@app.route('/')
def index():
    return 'OK', 200

@app.route('/generate-xlsx', methods=['POST'])
def generate_xlsx():
    data = request.get_json()

    estimate_name = data.get('estimateName', 'Estimate')
    customer_name = data.get('customerName', '')
    line_items = data.get('lineItems', [])
    totals = data.get('totals', {})

    wb = Workbook()
    ws = wb.active
    ws.title = 'Estimate'

    header_font = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
    header_fill = PatternFill(start_color='1E2127', end_color='1E2127', fill_type='solid')
    meta_label_font = Font(name='Calibri', color='888888', size=10)
    meta_font = Font(name='Calibri', bold=True, size=11)
    total_fill = PatternFill(start_color='E8F5C8', end_color='E8F5C8', fill_type='solid')
    thin_border = Border(bottom=Side(style='thin', color='DDDDDD'))
    group_fill = {
        'US':     PatternFill(start_color='EEF2FF', end_color='EEF2FF', fill_type='solid'),
        'CAN-EN': PatternFill(start_color='ECFDF5', end_color='ECFDF5', fill_type='solid'),
        'CAN-FR': PatternFill(start_color='FFF7ED', end_color='FFF7ED', fill_type='solid'),
    }

    # Build set of versions present in this estimate
    versions_present = set(item.get('version', '') for item in line_items)

    # Meta block
    ws['A1'] = 'Estimate'
    ws['A1'].font = Font(name='Calibri', bold=True, size=14)
    ws['B1'] = estimate_name
    ws['B1'].font = Font(name='Calibri', size=14)

    ws['A2'] = 'Customer'
    ws['A2'].font = meta_label_font
    ws['B2'] = customer_name
    ws['B2'].font = meta_font

    ws['A3'] = 'Total Packout'
    ws['A3'].font = meta_label_font
    ws['B3'] = safe_num(totals.get('totalPackout'))
    ws['B3'].font = Font(name='Calibri', bold=True, size=11, color='558B2F')
    ws['B3'].number_format = '"$"#,##0.00'

    meta_row = 2
    packout_map = [
        ('US',     'usPackout'),
        ('CAN-EN', 'canEnPackout'),
        ('CAN-FR', 'canFrPackout'),
    ]
    for label, key in packout_map:
        if label in versions_present:
            ws.cell(meta_row, 4).value = f'{label} Packout'
            ws.cell(meta_row, 4).font = meta_label_font
            ws.cell(meta_row, 5).value = safe_num(totals.get(key))
            ws.cell(meta_row, 5).number_format = '"$"#,##0.00'
            meta_row += 1

    ws.append([])
    ws.append([])

    col_headers = ['Version', 'Name', 'Width (in)', 'Height (in)', 'Quantity', 'Inking', 'Cutting', 'Unit Price', 'Total']
    header_row = 7
    ws.append(col_headers)
    for col_idx in range(1, len(col_headers) + 1):
        cell = ws.cell(row=header_row, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='left')

    ws.freeze_panes = 'A8'

    version_order = ['US', 'CAN-EN', 'CAN-FR']
    grouped = {v: [] for v in version_order}
    for item in line_items:
        v = item.get('version', '')
        if v in grouped:
            grouped[v].append(item)

    current_row = header_row + 1

    for version in version_order:
        items = grouped[version]
        if not items:
            continue

        fill = group_fill.get(version)

        for item in items:
            inking = ', '.join(item.get('inking', [])) if isinstance(item.get('inking'), list) else item.get('inking', '')
            cutting = ', '.join(item.get('cutting', [])) if isinstance(item.get('cutting'), list) else item.get('cutting', '')
            unit_price = round(safe_num(item.get('selectedCost') or item.get('estimatedUnitPrice')), 2)

            ws.append([
                item.get('version', ''),
                item.get('name', ''),
                item.get('trimWidth', 0),
                item.get('trimHeight', 0),
                item.get('quantity', 0),
                inking,
                cutting,
                unit_price,
                None
            ])

            for col in range(1, 10):
                cell = ws.cell(current_row, col)
                if fill:
                    cell.fill = fill
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left')

            ws.cell(current_row, 3).number_format = '0.000'
            ws.cell(current_row, 4).number_format = '0.000'
            ws.cell(current_row, 5).number_format = '#,##0'
            ws.cell(current_row, 8).number_format = '#,##0.00'
            ws.cell(current_row, 9).value = f'=E{current_row}*H{current_row}'
            ws.cell(current_row, 9).number_format = '#,##0.00'
            for col in [3, 4, 5, 8, 9]:
                ws.cell(current_row, col).alignment = Alignment(horizontal='right')

            current_row += 1

    # Total row
    ws.cell(current_row, 7).value = 'TOTAL PACKOUT'
    ws.cell(current_row, 7).font = Font(name='Calibri', bold=True, size=11)
    ws.cell(current_row, 9).value = safe_num(totals.get('totalPackout'))
    ws.cell(current_row, 9).number_format = '"$"#,##0.00'
    ws.cell(current_row, 9).font = Font(name='Calibri', bold=True, size=11, color='558B2F')
    for col in range(1, 10):
        ws.cell(current_row, col).fill = total_fill

    col_widths = [10, 55, 12, 12, 10, 12, 22, 12, 12]
    for i, width in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = width

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    encoded = base64.b64encode(buffer.read()).decode('utf-8')

    return jsonify({
        'file': encoded,
        'contentType': 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        'filename': f'{estimate_name}-{customer_name}.xlsx'
    })
